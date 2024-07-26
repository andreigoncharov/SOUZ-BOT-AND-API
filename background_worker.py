import asyncio
import datetime
import re

import smtplib
import time
import traceback
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import os

from aiogram import Bot
from openpyxl import Workbook
from openpyxl.styles import Alignment

from scripts.db_manager import RemoteDbManager
from scripts.config import *

import schedule

"""
Основные настройки
"""
smtp_user = 'andreigoncharov16072001@gmail.com'  # аккаунт с которого будут идти письма
smtp_password = 'bxxw wjhf slmo ukox'  # пароль аккаунта с которого будут идти письма

from_address = 'andreigoncharov16072001@gmail.com'  # адрес с которого отправлять
to_address = 'andreigoncharov1009@gmail.com'  # адрес куда отправлять

time_to_send = "20:10"  # время отправки в формате ЧЧ:ММ
"""
=========================================================
"""

loop = asyncio.get_event_loop()

rdb = RemoteDbManager()

smtp_server = 'smtp.gmail.com'
smtp_port = 587


async def send_email(file_name, date: str, isEmpty=False):
    msg = MIMEMultipart()
    msg['From'] = from_address
    msg['To'] = to_address
    msg['Subject'] = f'Отчет за {date}'
    if isEmpty:
        body = f'{date} не было отказов или отгрузок с корректировкой!'
        msg.attach(MIMEText(body, 'plain'))

    if not isEmpty:
        attachment = open(file_name, 'rb')
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(attachment.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f"attachment; filename= {os.path.basename(file_name)}")

        msg.attach(part)

    try:
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(smtp_user, smtp_password)
        text = msg.as_string()
        server.sendmail(from_address, to_address, text)
        server.quit()
    except Exception as e:
        print(f"Failed to send email: {e}")


async def create_excel_file(expeditors, expeditor_names, customer_names):
    wb = Workbook()
    ws = wb.active
    row_counter = 2
    merge_rows = []
    ws.append(["Клиент", "Документ", "Маршрут", "Номер в порядке отгрузки", "Статус"])
    for key, value in expeditors.items():
        ws.append([f"{re.sub(' +', ' ', str(expeditor_names.get(key)[0][0]).strip())} ({key})"])
        ws[f'A{row_counter}'].alignment = Alignment(horizontal='center', vertical='center')
        merge_rows.append(row_counter)
        row_counter += 1
        for order in value:
            ws.append(
                [f"{re.sub(' +', ' ', str(customer_names.get(order[0])[0][1]).strip())}", order[1], order[3], order[4],
                 order[2]])
            row_counter += 1
        ws.append([])
        row_counter += 1
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 2
        ws.column_dimensions[column].width = adjusted_width
    for index in merge_rows:
        ws.merge_cells(f'A{index}:C{index}')
    now = datetime.datetime.now()
    yesterday = now - datetime.timedelta(days=1)
    formatted_date = yesterday.strftime("%d-%m-%Y")
    wb.save(f"{formatted_date}.xlsx")


async def main():
    global rdb
    orders = await rdb.get_yesterday_docs(loop)
    now = datetime.datetime.now()
    yesterday = now - datetime.timedelta(days=1)
    formatted_date = yesterday.strftime("%d-%m-%Y")
    if len(orders) > 0:
        expeditors = {}
        expeditor_names = {}
        customers = []
        customer_names = {}
        for order in orders:
            if order[0] not in expeditors:
                expeditors[order[0]] = []
            if order[1] not in customers:
                customers.append(order[1])
            expeditors[order[0]].append([order[1], order[2], order[3]])
        for expeditor in expeditors:
            expeditor_names[expeditor] = await rdb.get_expeditor_name(expeditor, loop)

        for key, value in expeditors.items():
            new_arr = []
            for ord in value:
                route = await rdb.get_point_by_dock(ord[1], loop)
                route = route[0]
                # print(ord, ord[1], route)
                new_arr.append([ord[0], ord[1], ord[2], route[0], route[1]])
            expeditors[key] = new_arr
        for customer in customers:
            customer_names[customer] = await rdb.get_customer_description(customer, loop)
        await create_excel_file(expeditors, expeditor_names, customer_names)
        await send_email(f"{formatted_date}.xlsx", formatted_date)
    else:
        await send_email(f"{formatted_date}.xlsx", formatted_date, isEmpty=True)


def start():
    try:
        asyncio.run(main())
    except:
        try:
            asyncio.run(main())
        except:
            try:
                asyncio.run(main())
            except:
                bot = Bot(TOKEN)
                asyncio.run(bot.send_message(420404892, f"Ошибка отчета!\n \n {traceback.format_exc()}"))


# schedule.every().day.at(time_to_send).do(start)
#
# while True:
#     schedule.run_pending()
#     time.sleep(1)
start()